#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS分析レポート 自動生成スクリプト
====================================
使い方（KO形式）:
  python3 generate_report.py --xlsx 43345_KO.xlsx --csv 43345.csv --output report_43345.html

使い方（Clarityのみ）:
  python3 generate_report.py --csv 44402.csv --output report_44402.html

オプション:
  --xlsx      KOフォーマットのXLSXファイル（ファイル名自由）
  --csv       ClarityのCSVファイル（ファイル名自由）
  --image     LPページ画像ファイル（省略時はXLSX/CSVと同名を自動検索）
  --output    出力HTMLファイル名（デフォルト: report_output.html）
  --template  テンプレートHTMLパス（省略時は同フォルダのsms_report_template.html）
"""

import argparse
import csv
import json
import base64
import os
import re
import sys
from collections import defaultdict
import openpyxl

def find_data_file(script_dir, candidates, glob_pattern=None):
    """複数の候補ファイル名から存在するものを返す（日本語ファイル名の揺れ対応）"""
    import glob as _glob
    # glob で先に探す
    if glob_pattern:
        hits = _glob.glob(os.path.join(script_dir, glob_pattern))
        if hits:
            return hits[0]
    # 候補名リストで探す
    for name in candidates:
        p = os.path.join(script_dir, name)
        if os.path.exists(p):
            return p
    return None

# ── 離反期間バケット ──────────────────────────
SEG_ORDER = ['1ヶ月未満','1〜2ヶ月','2〜3ヶ月','3〜6ヶ月','半年〜1年','1〜3年','3年以上']

# カード使用率プレースホルダー（後で実績値に更新）
CARD_RATES = {
    '1ヶ月未満': 0.85,
    '1〜2ヶ月':  0.60,
    '2〜3ヶ月':  0.50,
    '3〜6ヶ月':  0.35,
    '半年〜1年': 0.20,
    '1〜3年':    0.10,
    '3年以上':   0.05,
}

# 来店結果報告の列マッピング（0始まり）
VISIT_RATE_COLS = {
    '1ヶ月未満': 2,
    '1〜2ヶ月':  3,
    '2〜3ヶ月':  4,
    '3〜6ヶ月':  5,
    '半年〜1年': 6,
    '1〜3年':    7,
    '3年以上':   8,
}

def load_visit_rates(xlsx_path):
    """来店結果報告_平滑化版.xlsxから効果測定日数別・離反期間別来店率を読み込む"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rates = {}  # {day: {segment_label: rate}}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None: continue
        try:
            day = int(row[0])
        except (TypeError, ValueError):
            continue
        day_rates = {}
        for label, col_idx in VISIT_RATE_COLS.items():
            val = row[col_idx]
            if val is not None:
                day_rates[label] = float(val)
        rates[day] = day_rates
    return rates

def get_segment(days):
    if days is None: return None
    if days < 30:    return '1ヶ月未満'
    if days < 60:    return '1〜2ヶ月'
    if days < 90:    return '2〜3ヶ月'
    if days < 180:   return '3〜6ヶ月'
    if days < 365:   return '半年〜1年'
    if days < 1095:  return '1〜3年'
    return '3年以上'

# ── SMSターゲット種別（最大離反期間・バースデー特別処理）────────
def load_campaign_targets(xlsx_path):
    """SMSターゲット.xlsx を読み込み、種別リストを返す
    戻り値: [{'name': str, 'max_seg': str, 'birthday': bool}, ...]
    """
    if not os.path.exists(xlsx_path):
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    result = []
    for row in rows[1:]:  # ヘッダー行スキップ
        if not row[0]: continue
        name    = str(row[0]).strip()
        max_seg = str(row[1]).strip() if row[1] else '1〜3年'
        note    = str(row[2]).strip() if row[2] else ''
        birthday = '12分の1' in note or 'バースデー' in name
        result.append({'name': name, 'max_seg': max_seg, 'birthday': birthday})
    return result

# ── 会員データ（台数別推定）────────────────────
MEMBER_DATA_SEG_MAP = {
    '1ヶ月未満':   '1ヶ月未満',
    '1〜2ヶ月未満': '1〜2ヶ月',
    '2〜3ヶ月未満': '2〜3ヶ月',
    '3～6ヶ月未満': '3〜6ヶ月',
    '半年～1年未満': '半年〜1年',
    '1年～3年未満': '1〜3年',
}

def load_member_estimates(xlsx_path, machines):
    """設置台数から各離反期間の推定会員数（排他的人数）を線形補間で算出。

    テーブルの人数列は累積値（その期間以内の全会員数）のため、
    排他的人数 = 当行の累積値 - 直前行の累積値 として計算する。
    例: 1〜2ヶ月のみの人数 = (2ヶ月未満累積) - (1ヶ月未満累積)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    lower = (machines // 100) * 100
    upper = lower + 100
    frac = (machines - lower) / 100

    # 台数→列インデックス: X台 → col = X/100 * 2
    lo_col = (lower // 100) * 2
    hi_col = (upper // 100) * 2

    # まず累積値を SEG_ORDER 順に収集
    cumulative = {}
    for row in rows[2:]:
        seg_key = row[1] if len(row) > 1 else None
        if seg_key not in MEMBER_DATA_SEG_MAP: continue
        label = MEMBER_DATA_SEG_MAP[seg_key]
        if label in cumulative: continue  # 最初の有効な行のみ（後半の#REF!行をスキップ）
        lo_val = row[lo_col] if len(row) > lo_col and row[lo_col] is not None else 0
        hi_val = row[hi_col] if len(row) > hi_col and row[hi_col] is not None else lo_val
        if not isinstance(lo_val, (int, float)): continue
        if not isinstance(hi_val, (int, float)): hi_val = lo_val
        cumulative[label] = round(lo_val + frac * (hi_val - lo_val))

    # 累積値→排他的人数（差分）に変換
    result = {}
    prev = 0
    for label in SEG_ORDER:
        if label == '3年以上': continue
        cum = cumulative.get(label, prev)
        result[label] = max(0, cum - prev)
        prev = cum
    return result

# ── 年代バケット ──────────────────────────────
AGE_ORDER = ['20代まで','30代','40代','50代','60代以上']

def get_age_segment(age):
    if age is None: return None
    if age <= 29: return '20代まで'
    if age <= 39: return '30代'
    if age <= 49: return '40代'
    if age <= 59: return '50代'
    return '60代以上'


# ── KO XLSX パーサー ─────────────────────────
def parse_ko_xlsx(filepath, visit_rate_data=None):
    """KOフォーマットのXLSXを読み込んでmeta・segments・KPIを返す"""
    try:
        import openpyxl
    except ImportError:
        print('openpyxlが必要です: pip install openpyxl --break-system-packages')
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    meta = {}

    # ── メタ情報（会員支持率詳細情報シート）
    ws1 = wb['会員支持率詳細情報']
    for row in ws1.iter_rows():
        for cell in row:
            v = cell.value
            if v is None: continue
            if v == '店舗名：':
                meta['store'] = ws1.cell(row=cell.row, column=cell.column+5).value
            elif v == '送信日時：':
                d = ws1.cell(row=cell.row, column=cell.column+6).value
                t = ws1.cell(row=cell.row, column=cell.column+16).value
                if d and t:
                    meta['datetime'] = f"{str(d)[:10]}T{str(t)}"
            elif v == '送信ID：':
                meta['sendId'] = str(int(ws1.cell(row=cell.row, column=cell.column+5).value))
            elif v == 'タイトル：':
                meta['campaign'] = ws1.cell(row=cell.row, column=cell.column+6).value
            elif v == '送信予約数':
                # 同じ列を下に探す（他列の値を誤取得しないよう列固定）
                col = cell.column
                for r in range(cell.row + 1, cell.row + 10):
                    val = ws1.cell(row=r, column=col).value
                    if isinstance(val, (int, float)) and val > 100:
                        meta['reserved'] = int(val)
                        break

    # ── SMS本文（K8 = 行8, 列11）
    sms_text = ws1.cell(row=8, column=11).value or ''
    meta['smsText'] = str(sms_text).strip()

    # ── 会員別データ集計（各会員様別送信詳細シート）
    ws2 = wb['各会員様別送信詳細']
    buckets = defaultdict(lambda: {'sent':0,'lp':0,'multi':0,'visits':0})
    total_sent = total_lp = total_multi = total_visits = 0
    max_last_visit = None  # 効果測定日数算出用

    for row in ws2.iter_rows(min_row=7, values_only=True):
        if row[0] is None: continue
        result   = row[3]
        lp_views = row[5] or 0
        days     = row[11]
        visited  = row[18]

        if result != '送達': continue
        total_sent += 1
        if lp_views >= 1: total_lp    += 1
        if lp_views >= 2: total_multi += 1
        if visited == '〇':
            total_visits += 1
            # col21（最終来店日 after sending）= row[20]
            last_visit_val = row[20]
            if last_visit_val:
                from datetime import date as _date
                if hasattr(last_visit_val, 'date'):
                    lv = last_visit_val.date()
                elif isinstance(last_visit_val, str):
                    try:
                        lv = _date.fromisoformat(last_visit_val[:10].replace('/', '-'))
                    except Exception:
                        lv = None
                else:
                    lv = None
                if lv and (max_last_visit is None or lv > max_last_visit):
                    max_last_visit = lv

        seg = get_segment(days)
        if seg is None: continue
        buckets[seg]['sent']   += 1
        if lp_views >= 1: buckets[seg]['lp']     += 1
        if lp_views >= 2: buckets[seg]['multi']  += 1
        if visited == '〇': buckets[seg]['visits'] += 1

    meta['sent']       = total_sent
    meta['lpViews']    = total_lp
    meta['multiViews'] = total_multi
    meta['visits']     = total_visits

    # 効果測定日数 = 送信日〜最終来店日（最大1〜）
    if max_last_visit and meta.get('datetime'):
        from datetime import date as _date
        send_dt = _date.fromisoformat(meta['datetime'][:10])
        measurement_days = max(1, (max_last_visit - send_dt).days)
    else:
        measurement_days = 0
    meta['measurementDays'] = measurement_days
    meta['pageViews']  = sum((row[5] or 0) for row in wb['各会員様別送信詳細'].iter_rows(min_row=7, values_only=True) if row[0] and row[3] == '送達')

    # ── 平均LP閲覧率・来店転換率（象限分類用）
    avg_lp_rate    = total_lp    / total_sent * 100 if total_sent else 0
    avg_visit_rate = total_visits / total_sent * 100 if total_sent else 0

    # ── 離反期間別セグメント配列生成
    segments = []
    for label in SEG_ORDER:
        d = buckets[label]
        if d['sent'] == 0: continue
        lp_rate    = round(d['lp']     / d['sent'] * 100, 1)
        visit_rate = round(d['visits'] / d['sent'] * 100, 1)
        # 来店結果報告データがあれば効果測定日数ベースの実績値を使用、なければフォールバック
        if visit_rate_data and meta.get('measurementDays'):
            ref_days = min(meta['measurementDays'], 30)
            card_rate = visit_rate_data.get(ref_days, {}).get(label, CARD_RATES.get(label, 0.5))
        else:
            card_rate = CARD_RATES.get(label, 0.5)

        # 象限分類（X=来店転換率、Y=LP閲覧率）
        high_visit = visit_rate >= avg_visit_rate
        high_lp    = lp_rate    >= avg_lp_rate
        if   high_visit and high_lp:    tag = 'q1'  # 優良反応層
        elif high_visit and not high_lp: tag = 'q2'  # 高来店層
        elif not high_visit and high_lp: tag = 'q3'  # 関心限定層
        else:                            tag = 'q4'  # 低反応層

        segments.append({
            'label':     label,
            'sent':      d['sent'],
            'lp':        d['lp'],
            'multi':     d['multi'],
            'visits':    d['visits'],
            'lpRate':    lp_rate,
            'visitRate': visit_rate,
            'cardRate':  card_rate,
            'tag':       tag,
        })

    # ── 年代別セグメント配列生成（4象限用）
    age_buckets = defaultdict(lambda: {'sent':0,'lp':0,'multi':0,'visits':0})
    wb2 = openpyxl.load_workbook(filepath)  # 再読み込み不要だが再利用
    ws2b = wb2['各会員様別送信詳細']
    for row in ws2b.iter_rows(min_row=7, values_only=True):
        if row[0] is None: continue
        if row[3] != '送達': continue
        age      = row[9]   # 年齢
        lp_views = row[5] or 0
        visited  = row[18]
        age_seg  = get_age_segment(age)
        if age_seg is None: continue
        age_buckets[age_seg]['sent']   += 1
        if lp_views >= 1: age_buckets[age_seg]['lp']     += 1
        if lp_views >= 2: age_buckets[age_seg]['multi']  += 1
        if visited == '〇': age_buckets[age_seg]['visits'] += 1

    age_segments = []
    for label in AGE_ORDER:
        d = age_buckets[label]
        if d['sent'] == 0: continue
        lp_rate    = round(d['lp']     / d['sent'] * 100, 1)
        visit_rate = round(d['visits'] / d['sent'] * 100, 1)
        high_visit = visit_rate >= avg_visit_rate
        high_lp    = lp_rate    >= avg_lp_rate
        if   high_visit and high_lp:     tag = 'q1'
        elif high_visit and not high_lp: tag = 'q2'
        elif not high_visit and high_lp: tag = 'q3'
        else:                            tag = 'q4'
        age_segments.append({
            'label':     label,
            'sent':      d['sent'],
            'lp':        d['lp'],
            'multi':     d['multi'],
            'visits':    d['visits'],
            'lpRate':    lp_rate,
            'visitRate': visit_rate,
            'tag':       tag,
        })

    return meta, segments, age_segments


# ── Clarity CSV パーサー ──────────────────────
def classify(pct):
    if pct >= 7.0:  return 'q1', 'keep'
    elif pct >= 3.0: return 'q2', 'fix'
    elif pct >= 1.5: return 'q3', 'fix'
    else:            return 'q4', 'del'

def parse_time(time_str):
    parts = time_str.strip().split(':')
    if len(parts) == 3:
        total = int(parts[0])*3600 + int(parts[1])*60 + int(parts[2])
        return f'{total//60}m{total%60}s' if total >= 60 else f'{total}s'
    return time_str

def parse_clarity_csv(filepath):
    scroll_depths = []
    meta = {}
    with open(filepath, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    for row in rows:
        if len(row) >= 2:
            if row[0] == 'プロジェクト名':    meta['project']   = row[1]
            elif row[0] == '閲覧済み URL を含む': meta['url']  = row[1]
            elif row[0] == 'ページ ビュー':   meta['pageViews'] = int(row[1])
    for row in rows:
        if len(row) >= 3 and row[0].strip().lstrip('"').isdigit():
            depth   = int(row[0].strip().strip('"'))
            time_raw = row[1].strip().strip('"')
            pct     = float(row[2].strip().strip('"').replace('%',''))
            tag, judge = classify(pct)
            scroll_depths.append({'depth':depth,'time':parse_time(time_raw),'pct':pct,'tag':tag,'judge':judge})
    return meta, scroll_depths

def detect_csv_type(filepath):
    fname = os.path.basename(filepath).lower()
    with open(filepath, encoding='utf-8-sig') as f:
        head = f.read(500)
    if 'メトリック' in head and 'Scroll' in head and 'ドロップ' in head:
        return 'scroll'
    if 'メトリック' in head and 'Attention' in head:
        return 'attention'
    if 'プロジェクト名' in head and 'スクロールの奥行き' in head:
        return 'attention'
    return 'unknown'

def parse_scroll_csv(filepath):
    """到達率CSV（Scroll metric）: スクロールの奥行き, 訪問者数, % ドロップ オフ"""
    reach_map = {}
    with open(filepath, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    max_visitors = None
    for row in rows:
        if len(row) >= 2 and row[0].strip().lstrip('"').isdigit():
            depth = int(row[0].strip().strip('"'))
            try:
                visitors = float(row[1].strip().strip('"'))
            except:
                continue
            if max_visitors is None:
                max_visitors = visitors
            reach = round(visitors / max_visitors * 100) if max_visitors else 100
            reach_map[depth] = reach
    return reach_map


# ── 画像変換 ──────────────────────────────────
def image_to_base64(filepath):
    ext  = os.path.splitext(filepath)[1].lower().lstrip('.')
    mime = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','webp':'image/webp'}.get(ext,'image/jpeg')
    with open(filepath,'rb') as f:
        return f'data:{mime};base64,{base64.b64encode(f.read()).decode()}'


# ── HTML注入 ──────────────────────────────────
def inject_scroll_depths(html, scroll_depths):
    new_str = 'scrollDepths: ' + json.dumps(scroll_depths, ensure_ascii=False, indent=4) + ','
    replacement = '  // LP スクロール深度データ（自動生成）\n  ' + new_str + '\n'
    pattern = r'// LP スクロール深度データ.*?scrollDepths: \[.*?\],\n'
    return re.sub(pattern, lambda m: replacement, html, flags=re.DOTALL)

def analyze_sms(text, store_name=''):
    """SMS本文の品質を簡易チェックして評価結果を返す"""
    if not text:
        return {'score': 'unknown', 'checks': [], 'charCount': 0}

    checks = []
    url_removed = re.sub(r'https?://\S+', '', text).strip()

    # ① 店名が入っているか（改行を除去して比較）
    text_flat = text.replace('\n', '').replace('\r', '')
    store_in_text = store_name and (store_name[:4] in text_flat or store_name[:6] in text_flat)
    checks.append({
        'label': '店名の記載',
        'status': 'ok' if store_in_text else 'warn',
        'detail': '店名が記載されており、どこからのSMSか明確です' if store_in_text
                  else '店名がないと迷惑SMSと勘違いされる可能性があります'
    })

    # ② お客様氏名（個人名差し込みのみ検出。汎用語「〇〇様」は除外）
    personal_name_patterns = ['{', '【氏名】', '[氏名]', '氏名', 'お名前', '○○様', '〇〇様', '◯◯様']
    has_personal_name = any(p in text for p in personal_name_patterns)
    # 非個人名の「様」表現を除去してから残った「様」で個人名を判定
    # 例: 「吉田健様\n会員様限定」→会員様を除去→「吉田健様」が残り個人名あり と判定
    if not has_personal_name:
        # ステップ1: 非個人名の「様」パターンを除去
        non_personal_sama = r'(?:お客|会員|皆|みなさ|ご来店|ご利用|ご参加|ご登録|来店|ご愛顧|御愛顧)様'
        text_cleaned = re.sub(non_personal_sama, '', text)
        # ステップ2: 「様」を含む複合語（模様・有様・神様・仏様）を除去
        text_cleaned = re.sub(r'模様|有様|神様|仏様', '', text_cleaned)
        # ステップ3: 直前が漢字/カタカナ(名前らしい文字)で直後が「々子式相な」でなければ個人名と判定
        # 「の様な」「した時の様な」は直前が平仮名なのでマッチしない
        has_personal_name = bool(re.search(r'[一-龯ァ-ヶ]{1,8}様(?![々子式相な])', text_cleaned))
    checks.append({
        'label': 'お客様名の記載',
        'status': 'ok' if has_personal_name else 'na',
        'detail': 'お客様名が差し込まれており承認欲求に働きかけています' if has_personal_name
                  else '個人名差し込みは確認できません（KOレポートからは判定不可）'
    })

    # ③ お店の思いが伝わるか（温かみ・感情語の有無）
    warm_words = ['ありがとう', 'おかげさま', 'うれし', '感謝', '喜', 'お待ち', 'ぜひ',
                  '皆様', 'みなさま', '心より', '特別', 'とっておき', '大切', '思い', '楽しみ']
    has_warmth = any(w in text for w in warm_words)
    content_len = len(re.sub(r'https?://\S+|\s|　', '', url_removed))
    too_short = content_len < 15
    checks.append({
        'label': 'お店の思い・温かみ',
        'status': 'ok' if (has_warmth and not too_short) else 'warn',
        'detail': 'お店の気持ちが感じられる文章です' if (has_warmth and not too_short)
                  else '作業的・短すぎる文章はお客様の興味を惹きにくい場合があります。感謝・期待感を添えると効果的です'
    })

    # ④ チラ見せ度（情報の出し過ぎ検出）
    text_no_url = re.sub(r'https?://\S+', '', text)
    has_date = bool(re.search(r'\d+月\d+日|\d+/\d+', text_no_url))
    event_keywords = ['リニューアル', 'オープン', 'OPEN', 'Open', '入替', '新台', 'フェア', 'グランド', '移転']
    has_event = any(kw in text_no_url for kw in event_keywords)
    tease_overdisclosed = has_date and has_event
    checks.append({
        'label': 'チラ見せ度',
        'status': 'warn' if tease_overdisclosed else 'ok',
        'detail': '日付とイベント内容が両方記載されており、LPを見る前に情報が出し過ぎている可能性があります。LP訪問の動機を残すよう内容を調整してください'
                  if tease_overdisclosed
                  else 'LPを見ないとわからない情報が残っており、クリック動機が保たれています'
    })

    # ⑤ 緊急性・限定感
    urgency_keywords = ['今月', '今週', 'GW', 'お正月', '年末', '年始', '今日', '明日', '今晩',
                        '期間限定', '今だけ', 'この機会', '数量限定']
    limit_keywords = ['VIP', '会員限定', '会員様だけ', '会員様限定', '特別な会員', '選ばれた',
                      '一部の会員', '限られた', '限定', 'プレミアム会員', 'ご優待']
    date_limit = bool(re.search(r'\d+月\d+日まで|\d+日間限定|\d+日まで', text_no_url))
    has_urgency = (any(kw in text for kw in urgency_keywords)
                   or any(kw in text for kw in limit_keywords)
                   or date_limit)
    checks.append({
        'label': '緊急性・限定感',
        'status': 'ok' if has_urgency else 'warn',
        'detail': '今行く理由・限定感がある訴求が含まれています' if has_urgency
                  else '期限や限定感がなく「今行かなくてもいい」と感じさせる可能性があります。「今月」「会員限定」などを加えると効果的です'
    })

    # ⑥ CTAの明確さ
    cta_keywords = ['↓', '▶', '▼', 'タップ', 'こちら', 'コチラ', '確認', '詳細', 'クリック',
                    '今すぐ', 'ご確認', 'チェック', 'ご覧', 'はこちら']
    has_cta = any(kw in text for kw in cta_keywords)
    checks.append({
        'label': 'CTAの明確さ',
        'status': 'ok' if has_cta else 'warn',
        'detail': 'LP誘導フレーズがあり、クリックを促す構成になっています' if has_cta
                  else 'URLだけの掲載はタップされにくくなります。「↓詳細はコチラ」「ご確認ください↓」など一言添えると効果的です'
    })

    # ⑦ 文章の使い回し感（自動判定不可・手動のみ）
    checks.append({
        'label': '文章の使い回し感',
        'status': 'na',
        'detail': '自動判定不可。毎回同じ文章を使い回している場合は手動でチェックしてください'
    })

    # 文字数（URL・改行含む）
    char_count = len(text)

    # ⑧ 文字数（情報表示のみ。長い・短いで優劣なし）
    if char_count <= 70:
        char_status, char_detail = 'info', f'{char_count}字：基本料金範囲内（70字以内）'
    elif char_count <= 133:
        char_status, char_detail = 'info', f'{char_count}字：推奨範囲内（133字以内）。追加料金あり'
    else:
        char_status, char_detail = 'warn', f'{char_count}字：133字を超えています。料金が上がる可能性があります'
    checks.append({
        'label': f'文字数 {char_count}字',
        'status': char_status,
        'detail': char_detail
    })

    warn_count = sum(1 for c in checks if c['status'] == 'warn')
    if warn_count >= 3:
        score = 'review'
    elif warn_count >= 1:
        score = 'caution'
    else:
        score = 'good'
    return {'score': score, 'checks': checks, 'charCount': char_count}


def analyze_heatmap_red_position(image_path):
    """LP画像のヒートマップから赤エリアの縦位置（%）を返す。
    赤ピクセル（R高・G低・B低）の重心をY軸方向で計算。
    検出できない場合は None を返す。
    """
    try:
        from PIL import Image
    except ImportError:
        return None
    try:
        img = Image.open(image_path).convert('RGB')
        w, h = img.size
        # 横幅を40pxに縮小して処理速度を上げる
        small = img.resize((40, h), Image.LANCZOS)
        pixels = small.load()
        row_weights = []
        for y in range(h):
            red_count = sum(
                1 for x in range(40)
                if pixels[x, y][0] > 160 and pixels[x, y][1] < 80 and pixels[x, y][2] < 80
            )
            row_weights.append(red_count)
        total = sum(row_weights)
        if total == 0:
            return None
        center_y = sum(y * row_weights[y] for y in range(h)) / total
        return round(center_y / h * 100)  # 0〜100の%
    except Exception:
        return None


def lp_actions_from_scroll(scroll_depths):
    """Scroll到達率データからLPの長さに関する改善提案を生成。
    到達率30%以下 かつ その深度が50%以下（残り半分以上放棄されている）の場合に提案。
    """
    if not scroll_depths:
        return []
    sorted_d = sorted(scroll_depths, key=lambda x: int(str(x['depth']).replace('%', '')))
    for d in sorted_d:
        depth_val = int(str(d['depth']).replace('%', ''))
        reach = d.get('reach')  # Scroll CSVから設定される到達率（なければスキップ）
        if reach is None:
            continue
        if reach <= 30 and depth_val <= 50:
            return [{
                'quad': 'q4',
                'title': '📏 LP長すぎ検出：短縮を推奨',
                'body': (f'スクロール深度{depth_val}%地点で到達率が{reach:.1f}%まで低下しており、'
                         f'LPの下半分以上がほぼ閲覧されていない状態です。'
                         f'重要な情報を上部に集約し、より短いLPへの変更を検討してください。'),
            }]
    return []


def lp_actions_from_heatmap(image_path, scroll_depths):
    """ヒートマップ赤エリア位置の到達率が70%未満の場合に上部移動を提案。"""
    if not image_path or not scroll_depths:
        return []
    red_pct = analyze_heatmap_red_position(image_path)
    if red_pct is None:
        return []
    # scroll_depthsの中で最も近い深度を探す
    sorted_d = sorted(scroll_depths, key=lambda x: int(str(x['depth']).replace('%', '')))
    best = min(sorted_d, key=lambda x: abs(int(str(x['depth']).replace('%', '')) - red_pct))
    reach = best.get('reach')  # Scroll CSVから設定される到達率
    depth_val = int(str(best['depth']).replace('%', ''))
    if reach is None or reach >= 70:
        return []
    return [{
            'quad': 'q3',
            'title': f'🔥 注目エリアの到達率が低下：上部移動を推奨',
            'body': (f'ヒートマップで最も注目されているエリア（LP上から約{red_pct}%付近）の'
                     f'スクロール到達率が{reach:.1f}%にとどまっています。'
                     f'このコンテンツをLPのより上部に移動することで、より多くのユーザーへのリーチが期待できます。'),
        }]
    return []


_SPECIAL_FEEL_KEYWORDS = {'限定', '特別', 'プレミアム', 'VIP', '特典', 'シークレット', '会員'}

def _sms_has_special_feel(meta):
    """SMS本文に既に特別感・限定感のキーワードが含まれるか判定"""
    text = meta.get('smsText', '') or ''
    return any(kw in text for kw in _SPECIAL_FEEL_KEYWORDS)


def generate_actions(segments, age_segments, meta, sms_analysis=None, extra_lp_actions=None, visit_rate_data=None):
    """実データから「次の一手」アクション提案リストを自動生成"""
    actions = []
    if not segments:
        return actions

    total_sent   = sum(s['sent']   for s in segments)
    total_visits = sum(s['visits'] for s in segments)
    total_lp     = sum(s['lp']     for s in segments)
    avg_visit_rate = round(total_visits / total_sent * 100, 1) if total_sent else 0
    avg_lp_rate    = round(total_lp    / total_sent * 100, 1) if total_sent else 0

    q1_segs = [s for s in segments if s['tag'] == 'q1']
    q3_segs = [s for s in segments if s['tag'] == 'q3']
    q4_segs = [s for s in segments if s['tag'] == 'q4']

    # ① 優良反応層（q1）
    if q1_segs:
        labels = '・'.join(s['label'] for s in q1_segs)
        total_q1 = sum(s['sent'] for s in q1_segs)
        best = max(q1_segs, key=lambda s: s['visitRate'])
        actions.append({
            'quad': 'q1',
            'title': f'★ 優良反応層（{labels}）：継続・追加配信',
            'body': (f'来店転換率・LP支持率ともに全体平均を上回る優良セグメント（計{total_q1}名）。'
                     f'最高来店転換率は{best["label"]}の{best["visitRate"]}%。'
                     f'次回も同条件で継続配信し、成功パターンを積み上げることを推奨。'),
        })

    # ② 関心限定層（q3）：LPは見るが来店しない
    if q3_segs:
        labels = '・'.join(s['label'] for s in q3_segs)
        q3_total = sum(s['sent'] for s in q3_segs)
        actions.append({
            'quad': 'q3',
            'title': f'△ 関心限定層（{labels}）：来店転換の底上げ',
            'body': (f'LP閲覧率は平均以上ながら来店転換率が低いセグメント（計{q3_total}名）。'
                     f'LP内容と店頭の一致度を高め、「今すぐ行く理由」を強化することで改善が期待できます。'),
        })

    # ③ 低反応層（q4）：最も来店率が低いセグメントを抽出（送信数20件未満は除外）
    if q4_segs:
        _q4_eligible = [s for s in q4_segs if s['sent'] >= 20] or q4_segs
        worst = min(_q4_eligible, key=lambda s: s['visitRate'])
        _low = worst['sent'] < 20
        _mdays = min(meta.get('measurementDays', 7), 30)
        _nat_val = visit_rate_data.get(_mdays, {}).get(worst['label']) if visit_rate_data else None
        _nat_str = f'（全国平均{round(_nat_val * 100, 1)}%）' if _nat_val is not None else ''
        _SHORT_LAPSE_SET  = {'1ヶ月未満', '1〜2ヶ月', '2〜3ヶ月'}
        _MEDIUM_LAPSE_SET = {'3〜6ヶ月', '半年〜1年未満'}
        if worst['label'] in _SHORT_LAPSE_SET:
            _q4_body = (
                f'{worst["sent"]}名送信に対しLP到達{worst["lp"]}名（{worst["lpRate"]}%）、'
                f'来店転換率{worst["visitRate"]}%{_nat_str}。'
                f'短期離反層は本来転換しやすい層のため、この結果は要注意。'
                f'SMS訴求内容の見直しを優先的に検討してください。'
            )
        elif worst['label'] in _MEDIUM_LAPSE_SET:
            if _sms_has_special_feel(meta):
                _q4_body = (
                    f'{worst["sent"]}名送信に対しLP到達{worst["lp"]}名（{worst["lpRate"]}%）、'
                    f'来店転換率{worst["visitRate"]}%{_nat_str}。'
                    f'今回のSMS本文には既に限定感のある訴求が含まれています。'
                    f'SMS訴求の方向性は維持しつつ、LP内の来店を促す訴求をより具体的・魅力的に強化することを優先的に検討してください。'
                )
            else:
                _q4_body = (
                    f'{worst["sent"]}名送信に対しLP到達{worst["lp"]}名（{worst["lpRate"]}%）、'
                    f'来店転換率{worst["visitRate"]}%{_nat_str}。'
                    f'中期離反層への再来店には、通常配信より特別感・限定感のある訴求が必要です。'
                    f'「会員限定」「期間限定」など特別感を前面に出したSMS本文とLPへの改善を検討してください。'
                )
        else:
            if _sms_has_special_feel(meta):
                _q4_body = (
                    f'{worst["sent"]}名送信に対しLP到達{worst["lp"]}名（{worst["lpRate"]}%）、'
                    f'来店転換率{worst["visitRate"]}%{_nat_str}。'
                    f'今回のSMS本文には既に限定感のある訴求が含まれています。'
                    f'長期離反層への配信は内容の質が成否を左右するため、LP内の来店を促す訴求をより具体的・魅力的に強化することを優先してください。'
                )
            else:
                _q4_body = (
                    f'{worst["sent"]}名送信に対しLP到達{worst["lp"]}名（{worst["lpRate"]}%）、'
                    f'来店転換率{worst["visitRate"]}%{_nat_str}。'
                    f'長期離反層への配信は特別なタイミングに絞るからこそ、訴求内容の質が成否を左右します。'
                    f'「会員限定」「期間限定」など特別感・限定感を前面に出したSMS本文とLPへの改善を優先してください。'
                )
        actions.append({
            'quad': 'q4',
            'title': f'▽ 低反応層 › {worst["label"]}離反（{worst["sent"]}名）：要優先対応',
            'body': _q4_body,
            'lowSample': _low,
            'sentCount': worst['sent'],
        })

    # ④ SMS品質：要改善（warn）が2つ以上の場合に本文見直し提案
    if sms_analysis:
        # 文字数（info）を除く5項目のみ対象
        target_checks = [c for c in sms_analysis.get('checks', [])
                         if c['status'] != 'info']
        warn_items = [c['label'] for c in target_checks if c['status'] == 'warn']
        na_items   = [c['label'] for c in target_checks if c['status'] == 'na']
        # warnが2つ以上、またはwarn+naが合計2つ以上の場合に提案
        if len(warn_items) >= 2 or (len(warn_items) + len(na_items)) >= 2:
            problem_labels = warn_items + na_items
            # 各項目の改善ヒントを生成
            hints = []
            hint_map = {
                '店名の記載':      '店名を必ず冒頭に入れ、迷惑SMS誤認を防ぎましょう',
                'お客様名の記載':  '個人名差し込みで開封率・反応率が上がります',
                'お店の思い・温かみ': '感謝や期待感を一言添えると読み手の心に刺さります',
                '汎用フレーズのみ': '「何があるの？」と思わせる具体的なキーワードを追加しましょう',
                '興味喚起フック':  '数字・限定・固有ワードで「気になる」を演出してください',
            }
            for label in problem_labels:
                if label in hint_map:
                    hints.append(f'・{label}：{hint_map[label]}')
            body = (f'SMS本文チェックで{len(problem_labels)}項目の要改善・要確認が検出されました。'
                    f'次回配信前に本文の見直しを推奨します。<br>' + '<br>'.join(hints))
            actions.append({
                'quad': 'q4',
                'title': f'✏️ SMS本文の見直し提案（{len(problem_labels)}項目要対応）',
                'body': body,
            })

    # ⑥ LP分析からの追加提案（スクロール到達率・ヒートマップ）
    if extra_lp_actions:
        actions.extend(extra_lp_actions)

    return actions


def generate_findings(segments, age_segments, meta, sms_analysis=None, visit_rate_data=None):
    """実データから所見・コメントテキストを自動生成"""
    if not segments:
        return '（データが不足しているため所見を生成できません）'

    total_sent   = sum(s['sent']   for s in segments)
    total_visits = sum(s['visits'] for s in segments)
    total_lp     = sum(s['lp']     for s in segments)
    avg_visit_rate = round(total_visits / total_sent * 100, 1) if total_sent else 0
    avg_lp_rate    = round(total_lp    / total_sent * 100, 1) if total_sent else 0

    best_seg  = max(segments, key=lambda s: s['visitRate'])
    _MIN_SENT = 20
    _eligible = [s for s in segments if s['sent'] >= _MIN_SENT] or segments
    worst_seg = min(_eligible, key=lambda s: s['visitRate'])

    store = meta.get('store', '今回の店舗')
    lines = []

    # ── 全体総括
    lines.append('【全体総括】')
    if age_segments:
        # q1（優良反応層）の年代を全て列挙。q1がなければ最高visitRateの1つ
        q1_ages = [a for a in age_segments if a.get('tag') == 'q1']
        if not q1_ages:
            q1_ages = [max(age_segments, key=lambda a: a['visitRate'])]
        age_labels = '・'.join(a['label'] for a in q1_ages)
        lines.append(
            f'今回の配信では{age_labels}が優良反応層に位置し、'
            f'来店転換率・LP支持率ともに全体平均を上回る結果となった。'
        )
    else:
        lines.append(
            f'{store}の今回の配信では、{best_seg["label"]}離反層が最も高い来店転換率'
            f'（{best_seg["visitRate"]}%）を示した。'
        )

    lines.append('')

    # ── LP支持率
    lines.append('【LP支持率について】')
    if age_segments and len(age_segments) >= 2:
        max_lp_age = max(age_segments, key=lambda a: a['lpRate'])
        min_lp_age = min(age_segments, key=lambda a: a['lpRate'])
        diff = round(max_lp_age['lpRate'] - min_lp_age['lpRate'], 1)
        if diff > 8:
            lines.append(
                f'LP支持率はSMS本文の件名・冒頭一文・CTAの訴求力に大きく左右される。'
                f'{max_lp_age["label"]}のLP支持率（{max_lp_age["lpRate"]}%）に対して'
                f'{min_lp_age["label"]}（{min_lp_age["lpRate"]}%）に{diff}ポイントの差が見られることから、'
                f'年代別にSMSメッセージを最適化することで支持率の底上げが期待できる。'
            )
        else:
            lines.append(f'全体LP支持率は{avg_lp_rate}%で年代間の差は比較的小さく、安定した訴求ができている。')
    else:
        lines.append(f'全体LP支持率は{avg_lp_rate}%。SMS本文の件名・冒頭一文の最適化でさらなる改善が見込める。')

    lines.append('')

    # ── 低反応層への対応
    _SHORT_LAPSE  = {'1ヶ月未満', '1〜2ヶ月', '2〜3ヶ月'}
    _MEDIUM_LAPSE = {'3〜6ヶ月', '半年〜1年未満'}
    _fmdays = min(meta.get('measurementDays', 7), 30)
    if worst_seg['visitRate'] < avg_visit_rate:
        lines.append('【低反応層への対応】')
        _wlabel = worst_seg['label']
        _wvr    = worst_seg['visitRate']
        _wlpr   = worst_seg['lpRate']
        _fnat   = visit_rate_data.get(_fmdays, {}).get(_wlabel) if visit_rate_data else None
        _fnat_str = f'（全国平均{round(_fnat * 100, 1)}%）' if _fnat is not None else ''
        if _wlabel in _SHORT_LAPSE:
            lines.append(
                f'{_wlabel}離反層は来店転換率{_wvr}%{_fnat_str}。'
                f'本来再来店ハードルが低い短期離反層での低転換率は要注意。'
                f'SMS訴求内容の見直しを優先的に検討してください。'
            )
        elif _wlabel in _MEDIUM_LAPSE:
            if _sms_has_special_feel(meta):
                lines.append(
                    f'{_wlabel}離反層は来店転換率{_wvr}%{_fnat_str}。'
                    f'今回のSMS本文には既に限定感のある訴求が含まれています。'
                    f'SMS訴求の方向性は維持しつつ、LP内の来店を促す訴求をより具体的・魅力的に強化することを優先的に検討してください。'
                )
            else:
                lines.append(
                    f'{_wlabel}離反層は来店転換率{_wvr}%{_fnat_str}。'
                    f'中期離反層への再来店には、通常配信より特別感・限定感のある訴求が必要です。'
                    f'「会員限定」「期間限定」など特別感を前面に出したSMS本文とLPへの改善を検討してください。'
                )
        else:
            # 長期離反：低転換率は想定内。特別感ある訴求とLP改善が鍵
            if _sms_has_special_feel(meta):
                lines.append(
                    f'{_wlabel}離反層は来店転換率{_wvr}%{_fnat_str}。離反期間が長いほど転換率が低下する傾向は全国的に見られ、今回の結果は想定の範囲内。'
                    f'今回のSMS本文には既に限定感のある訴求が含まれています。'
                    f'長期離反層への配信は内容の質が成否を左右するため、LP内の来店を促す訴求をより具体的・魅力的に強化することを優先してください。'
                )
            else:
                lines.append(
                    f'{_wlabel}離反層は来店転換率{_wvr}%{_fnat_str}。離反期間が長いほど転換率が低下する傾向は全国的に見られ、今回の結果は想定の範囲内。'
                    f'長期離反層への配信は特別なタイミングに絞るからこそ、訴求内容の質が成否を左右します。'
                    f'「会員限定」「期間限定」など特別感・限定感を前面に出したSMS本文とLPへの改善を優先してください。'
                )
        lines.append('')

    # ── 次回提案
    lines.append('【次回提案】')
    lines.append('')

    # ◆ 離反期間別
    lines.append('◆ 離反期間別')
    _best_diff   = round(best_seg['visitRate'] - avg_visit_rate, 1)
    _best_vs     = f'全体平均より{_best_diff}pt高い' if _best_diff >= 0 else f'全体平均より{abs(_best_diff)}pt低い'
    _best_nat    = visit_rate_data.get(_fmdays, {}).get(best_seg['label']) if visit_rate_data else None
    _best_nat_str = f'・全国平均{round(_best_nat * 100, 1)}%' if _best_nat is not None else ''
    lines.append(
        f'・{best_seg["label"]}離反層（来店転換率{best_seg["visitRate"]}%、{_best_vs}{_best_nat_str}）：'
        f'好反応を維持。次回も中心ターゲットとして継続配信を推奨します。'
    )
    if worst_seg['label'] != best_seg['label']:
        _worst_diff   = round(worst_seg['visitRate'] - avg_visit_rate, 1)
        _worst_vs     = f'全体平均より{abs(_worst_diff)}pt低い' if _worst_diff < 0 else f'全体平均と同水準'
        _worst_nat    = visit_rate_data.get(_fmdays, {}).get(worst_seg['label']) if visit_rate_data else None
        _worst_nat_str = f'・全国平均{round(_worst_nat * 100, 1)}%' if _worst_nat is not None else ''
        if worst_seg['label'] in _SHORT_LAPSE:
            _w_action = 'SMS訴求内容の早急な見直しを優先してください。'
        elif worst_seg['label'] in _MEDIUM_LAPSE:
            if _sms_has_special_feel(meta):
                _w_action = 'LP内の来店を促す訴求をより具体的・魅力的に強化することを優先してください。'
            else:
                _w_action = '特別感・限定感のある本文・LP内容への改善を検討してください。'
        else:
            _w_action = '特別なタイミングに絞った配信で、訴求内容の質向上を優先してください。'
        lines.append(
            f'・{worst_seg["label"]}離反層（来店転換率{worst_seg["visitRate"]}%、{_worst_vs}{_worst_nat_str}）：{_w_action}'
        )
    lines.append('・ターゲット条件の緩和（離反期間の段階的な拡大）も引き続き検討してください。')
    lines.append('')

    # ◆ 年代別
    if age_segments:
        lines.append('◆ 年代別')
        q1_ages = [a for a in age_segments if a.get('tag') == 'q1']
        q2_ages = [a for a in age_segments if a.get('tag') == 'q2']
        q3_ages = [a for a in age_segments if a.get('tag') == 'q3']
        q4_ages = [a for a in age_segments if a.get('tag') == 'q4']
        for a in q1_ages:
            lines.append(
                f'・{a["label"]}（来店転換率{a["visitRate"]}%・LP支持率{a["lpRate"]}%、ともに全体平均以上）：'
                f'優良反応層。引き続き継続配信を推奨します。'
            )
        for a in q2_ages:
            lines.append(
                f'・{a["label"]}（来店転換率{a["visitRate"]}%は全体平均以上、LP支持率{a["lpRate"]}%は平均以下）：'
                f'SMSの件名・冒頭文を改善してLP到達率を底上げすることで、さらなる来店増が見込めます。'
            )
        for a in q3_ages:
            lines.append(
                f'・{a["label"]}（LP支持率{a["lpRate"]}%は全体平均以上、来店転換率{a["visitRate"]}%は平均以下）：'
                f'LPへの関心はあるが来店に繋がっていません。LP内の来店を促す訴求の具体化・強化を検討してください。'
            )
        for a in q4_ages:
            if _sms_has_special_feel(meta):
                _q4_text = 'LP内容の改善を優先してください。この年代に響くLP訴求への具体的な見直しを検討してください。'
            else:
                _q4_text = 'SMS本文とLP両方の改善が必要です。この年代に響く訴求内容への見直しを検討してください。'
            lines.append(
                f'・{a["label"]}（来店転換率{a["visitRate"]}%・LP支持率{a["lpRate"]}%、ともに全体平均以下）：'
                f'{_q4_text}'
            )
        lines.append('')

    if sms_analysis and sms_analysis.get('score') in ('caution', 'review'):
        lines.append('◆ SMS品質')
        lines.append('・SMS本文の品質改善（要確認項目の解消）')

    return '\n'.join(lines)


def inject_actions(html, actions):
    """DATAオブジェクトのactionsを実データに差し替え"""
    new_str = 'actions: ' + json.dumps(actions, ensure_ascii=False, indent=4) + ','
    pattern = r'actions: \[.*?\],'
    # lambdaを使うことでre.subによる\nの展開を防ぐ
    return re.sub(pattern, lambda m: new_str, html, flags=re.DOTALL)


def inject_findings(html, findings_text):
    """所見・コメントのプレースホルダーを実データに差し替え"""
    return html.replace('<!-- MEMO_CONTENT_PLACEHOLDER -->', findings_text)


def inject_sms_analysis(html, sms_text, analysis):
    """SMS本文と分析結果をDATAオブジェクトのプレースホルダーに注入"""
    # json.dumpsで確実にエスケープ
    text_json   = json.dumps(sms_text, ensure_ascii=False)
    checks_json = json.dumps(analysis['checks'], ensure_ascii=False)
    # プレースホルダーを実データで置換
    html = html.replace(
        "  smsText: '',\n  smsScore: 'unknown',\n  smsChecks: [],",
        f"  smsText: {text_json},\n  smsScore: '{analysis['score']}',\n  smsChecks: {checks_json},"
    )
    return html


def inject_age_segments(html, age_segments):
    new_str = 'ageSegments: ' + json.dumps(age_segments, ensure_ascii=False, indent=4) + ','
    replacement = '  // 年代別データ（自動生成）\n  ' + new_str + '\n'
    pattern = r'// 年代別データ.*?ageSegments: \[.*?\],\n'
    result = re.sub(pattern, lambda m: replacement, html, flags=re.DOTALL)
    if result == html:  # パターンが見つからない場合はsegmentsの直後に挿入
        html = html.replace('  // 離反期間別データ（自動生成）', replacement + '\n  // 離反期間別データ（自動生成）')
        return html
    return result

def inject_segments(html, segments):
    new_str = 'segments: ' + json.dumps(segments, ensure_ascii=False, indent=4) + ','
    replacement = '  // 離反期間別データ（自動生成）\n  ' + new_str + '\n'
    pattern = r'// 離反期間別データ.*?segments: \[.*?\],\n'
    return re.sub(pattern, lambda m: replacement, html, flags=re.DOTALL)

def inject_unsent_segments(html, segments, member_estimates, visit_rate_data, measurement_days,
                           max_segment=None, birthday_mode=False):
    """設置台数ベースの推定会員数と送信済み数から未送信層を計算してDATAに注入。
    max_segment: この離反期間を超えるセグメントはシミュレーションに含めない
    birthday_mode: True の場合、追加対象人数を 1/12 に補正（誕生日月限定施策）
    """
    ref_days = min(measurement_days, 30) if measurement_days else None

    # 最大対象期間のインデックスを決定
    max_idx = len(SEG_ORDER) - 1  # デフォルト：全セグメント
    if max_segment and max_segment in SEG_ORDER:
        max_idx = SEG_ORDER.index(max_segment)

    unsent = []
    for i, label in enumerate(SEG_ORDER):
        if label == '3年以上': continue
        if i > max_idx: continue  # キャンペーン種別の最大対象期間を超えたらスキップ
        est_total = member_estimates.get(label, 0)
        seg = next((s for s in segments if s['label'] == label), None)
        sent_count = seg['sent'] if seg else 0
        unsent_count = max(0, est_total - sent_count)
        if unsent_count == 0: continue
        # バースデー施策：誕生日月のお客様のみ対象 → 1/12 に補正
        if birthday_mode:
            unsent_count = max(1, unsent_count // 12)
        if visit_rate_data and ref_days:
            card_rate = visit_rate_data.get(ref_days, {}).get(label, CARD_RATES.get(label, 0.1))
        else:
            card_rate = CARD_RATES.get(label, 0.1)
        lp_rate = round(seg['lpRate'] / 100, 3) if seg else 0.25
        unsent.append({
            'label':       label,
            'estTotal':    est_total,
            'sentCount':   sent_count,
            'unsentCount': unsent_count,
            'lpRate':      lp_rate,
            'cardRate':    round(card_rate, 4),
        })
    new_str = 'unsentSegments: ' + json.dumps(unsent, ensure_ascii=False, indent=4) + ','
    pattern = r'unsentSegments:\s*\[.*?\],\n'
    return re.sub(pattern, new_str + '\n', html, flags=re.DOTALL)

def inject_kpi(html, meta):
    # KPIブロック全体を一括置換（行頭の2スペースインデントで限定）
    fields = {
        'reserved':        meta.get('reserved',        0),
        'sent':            meta.get('sent',             0),
        'lpViews':         meta.get('lpViews',          0),
        'multiViews':      meta.get('multiViews',       0),
        'totalPV':         meta.get('pageViews',        0),
        'visits':          meta.get('visits',           0),
        'measurementDays': meta.get('measurementDays',  0),
        'machines':        meta.get('machines',          0),
    }
    for key, val in fields.items():
        # 行頭2スペース + キー名: でKPIフィールドだけをマッチ
        html = re.sub(rf'(\n  {key}:\s*)\d+', rf'\g<1>{val}', html, count=1)
    return html

def inject_header(html, store=None, campaign=None, sendid=None, dt=None, axis=None):
    if store:    html = re.sub(r"store:\s*'[^']*'",    f"store: '{store}'",       html)
    if campaign: html = re.sub(r"campaign:\s*'[^']*'", f"campaign: '{campaign}'", html)
    if sendid:   html = re.sub(r"sendId:\s*'[^']*'",   f"sendId: '{sendid}'",     html)
    if dt:       html = re.sub(r"datetime:\s*'[^']*'", f"datetime: '{dt}'",       html)
    if axis:     html = re.sub(r"axis:\s*'[^']*'",     f"axis: '{axis}'",         html)
    return html

def inject_image(html, image_src):
    return re.sub(r'<img src="data:image/[^"]*" class="heatmap-img"',
                  f'<img src="{image_src}" class="heatmap-img"', html)

def inject_campaign_meta(html, campaign_type, max_segment):
    """campaignType / maxSegment を DATA に注入"""
    if campaign_type:
        html = re.sub(r"campaignType:\s*'[^']*'", f"campaignType: '{campaign_type}'", html)
    if max_segment:
        html = re.sub(r"maxSegment:\s*'[^']*'",   f"maxSegment: '{max_segment}'",     html)
    return html

def find_image(base_path):
    for ext in ['.jpg','.jpeg','.png','.webp']:
        p = base_path + ext
        if os.path.exists(p): return p
    return None


# ── コア処理（CLIもStreamlitも共通） ─────────────────────
def generate_report_core(
    xlsx_path=None,
    scroll_csv_path=None,
    attention_csv_path=None,
    image_path=None,
    machines=None,
    campaign_type=None,
    template_path=None,
    script_dir=None,
    store_name_status=None,     # '有'→ok / '無'→warn / None（自動判定）
    customer_name_status=None,  # '有'→ok / '無'→na  / None（自動判定）
    warmth_status=None,         # '有'→ok / '無'→warn / None（自動判定）
    tease_status=None,          # '無'→ok（出し過ぎなし）/ '有'→warn（出し過ぎ）/ None（自動判定）
    urgency_status=None,        # '有'→ok / '無'→warn / None（自動判定）
    cta_status=None,            # '有'→ok / '無'→warn / None（自動判定）
    reuse_status=None,          # '無'→ok（使い回しなし）/ '有'→warn（使い回しあり）/ None（判定不可）
    measurement_days=None,      # 手動設定時の想定効果測定日数（int）/ None のとき自動計算
):
    """
    レポートHTMLを生成して文字列で返す。
    script_dir を指定しない場合はこのファイルのディレクトリを使用。
    """
    if script_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    # テンプレート読み込み
    tpl = template_path or os.path.join(script_dir, 'sms_report_template.html')
    with open(tpl, encoding='utf-8') as f:
        html = f.read()

    meta = {}
    segments = []
    scroll_depths = []

    # 来店結果報告データ
    visit_rate_data = None
    rate_file = find_data_file(script_dir, ['visit_rates.xlsx', '来店結果報告_平滑化版.xlsx', '報告結果_平滑化版.xlsx'], glob_pattern='*報告*平滑*.xlsx')
    if rate_file:
        visit_rate_data = load_visit_rates(rate_file)

    # XLSX（KO形式）
    age_segments = []
    if xlsx_path:
        meta, segments, age_segments = parse_ko_xlsx(xlsx_path, visit_rate_data=visit_rate_data)
        html = inject_segments(html, segments)
        html = inject_age_segments(html, age_segments)

    # 想定効果測定日数の手動上書き
    if measurement_days is not None:
        meta['measurementDays'] = int(measurement_days)

    # 設置台数・キャンペーン種別
    target_file = find_data_file(script_dir, ['SMSターゲット.xlsx', 'SMS対象.xlsx', 'sms_targets.xlsx'])
    member_file = find_data_file(script_dir, ['member_data.xlsx', '会員データ調査_想定台数100刻み_算出済.xlsx', '会員データ調査_想定人数100刻み_算出済.xlsx'], glob_pattern='会員データ調査*.xlsx')
    campaign_targets = load_campaign_targets(target_file) if target_file else []

    max_segment   = None
    birthday_mode = False
    if campaign_type:
        matched = next((ct for ct in campaign_targets if ct['name'] == campaign_type), None)
        if matched:
            max_segment   = matched['max_seg']
            birthday_mode = matched['birthday']

    if machines:
        meta['machines'] = machines
    if machines and xlsx_path and member_file and os.path.exists(member_file):
        member_estimates = load_member_estimates(member_file, machines)
        html = inject_unsent_segments(html, segments, member_estimates, visit_rate_data,
                                      meta.get('measurementDays', 0),
                                      max_segment=max_segment, birthday_mode=birthday_mode)
        if campaign_type:
            html = inject_campaign_meta(html, campaign_type, max_segment or '')

    # Scroll CSV（到達率）
    reach_map = {}
    if scroll_csv_path:
        reach_map = parse_scroll_csv(scroll_csv_path)

    # Attention CSV（注目割合）
    if attention_csv_path:
        csv_type = detect_csv_type(attention_csv_path)
        if csv_type in ('attention', 'clarity'):
            clarity_meta, scroll_depths = parse_clarity_csv(attention_csv_path)
            meta.update({k: v for k, v in clarity_meta.items() if k not in meta})
            if not meta.get('sendId') and clarity_meta.get('url'):
                m = re.search(r'/(\d+)$', clarity_meta['url'])
                if m: meta['sendId'] = m.group(1)
            if reach_map:
                for d in scroll_depths:
                    d['reach'] = reach_map.get(d['depth'], None)
            html = inject_scroll_depths(html, scroll_depths)
    elif reach_map:
        # Scroll のみの場合も到達率データを注入
        html = inject_scroll_depths(html, [])

    # SMS本文分析
    if xlsx_path and meta.get('smsText'):
        sms_analysis = analyze_sms(meta['smsText'], store_name=meta.get('store', ''))
        # 手動選択値でoverride
        _overrides = {
            '店名の記載': (store_name_status, {
                '有': ('ok',   '店名が記載されており、どこからのSMSか明確です'),
                '無': ('warn', '店名がないと迷惑SMSと勘違いされる可能性があります'),
            }),
            'お客様名の記載': (customer_name_status, {
                '有': ('ok',   'お客様名が差し込まれており承認欲求に働きかけています'),
                '無': ('na',   '個人名差し込みは確認できません（KOレポートからは判定不可）'),
            }),
            'お店の思い・温かみ': (warmth_status, {
                '有': ('ok',   'お店の気持ちが感じられる文章です'),
                '無': ('warn', '作業的・短すぎる文章はお客様の興味を惹きにくい場合があります'),
            }),
            'チラ見せ度': (tease_status, {
                '無': ('ok',   'LPを見ないとわからない情報が残っており、クリック動機が保たれています'),
                '有': ('warn', '日付とイベント内容が両方記載されており、LPを見る前に情報が出し過ぎている可能性があります'),
            }),
            '緊急性・限定感': (urgency_status, {
                '有': ('ok',   '今行く理由・限定感がある訴求が含まれています'),
                '無': ('warn', '期限や限定感がなく「今行かなくてもいい」と感じさせる可能性があります'),
            }),
            'CTAの明確さ': (cta_status, {
                '有': ('ok',   'LP誘導フレーズがあり、クリックを促す構成になっています'),
                '無': ('warn', 'URLだけの掲載はタップされにくくなります。「↓詳細はコチラ」など一言添えると効果的です'),
            }),
            '文章の使い回し感': (reuse_status, {
                '無': ('ok',   '配信ごとに独自の内容が入っており、マンネリ化していません'),
                '有': ('warn', '同じ文章の使い回しは飽きにつながります。配信ごとに内容を変えることを推奨します'),
            }),
        }
        for check in sms_analysis['checks']:
            override_val, status_map = _overrides.get(check['label'], (None, {}))
            if override_val is not None and override_val in status_map:
                check['status'], check['detail'] = status_map[override_val]
        # スコア再計算（naはカウントしない）
        warn_count = sum(1 for c in sms_analysis['checks'] if c['status'] == 'warn')
        sms_analysis['score'] = 'review' if warn_count >= 3 else ('caution' if warn_count >= 1 else 'good')
        html = inject_sms_analysis(html, meta['smsText'], sms_analysis)

    # 次の一手・所見コメント（自動生成）
    if xlsx_path and segments:
        sms_analysis_ref = sms_analysis if (xlsx_path and meta.get('smsText')) else None
        # LP分析からの追加提案
        extra_lp = []
        extra_lp.extend(lp_actions_from_scroll(scroll_depths))
        extra_lp.extend(lp_actions_from_heatmap(image_path, scroll_depths))
        actions  = generate_actions(segments, age_segments, meta, sms_analysis_ref,
                                    extra_lp_actions=extra_lp, visit_rate_data=visit_rate_data)
        findings = generate_findings(segments, age_segments, meta, sms_analysis_ref,
                                     visit_rate_data=visit_rate_data)
        html = inject_actions(html, actions)
        html = inject_findings(html, findings)

    # KPI・ヘッダー
    if xlsx_path:
        html = inject_kpi(html, meta)
        html = inject_header(html,
            store    = meta.get('store'),
            campaign = meta.get('campaign'),
            sendid   = meta.get('sendId'),
            dt       = meta.get('datetime'),
            axis     = '来店転換率 × LP支持率',
        )
    elif attention_csv_path:
        html = inject_header(html, sendid=meta.get('sendId'))

    # 画像
    if image_path:
        html = inject_image(html, image_to_base64(image_path))

    return html


# ── メイン ────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='SMS分析レポート自動生成')
    parser.add_argument('--xlsx',       default=None, help='KO XLSXファイル')
    parser.add_argument('--csv',        default=None, help='Clarity注目割合CSVファイル（Attention）')
    parser.add_argument('--scroll-csv', default=None, help='Clarityスクロール深度CSVファイル（Scroll）')
    parser.add_argument('--image',      default=None, help='LPページ画像ファイル')
    parser.add_argument('--machines',      default=None, type=int, help='店舗の設置台数（緩和シミュレーション用）')
    parser.add_argument('--campaign-type', default=None, help='SMSキャンペーン種別（未指定時はインタラクティブ選択）')
    parser.add_argument('--template', default=None, help='テンプレートHTMLパス')
    parser.add_argument('--output',   default='report_output.html', help='出力ファイル名')
    args = parser.parse_args()

    if not args.xlsx and not args.csv:
        print('エラー: --xlsx または --csv を指定してください')
        sys.exit(1)

    # テンプレート読み込み
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = args.template or os.path.join(script_dir, 'sms_report_template.html')
    if not os.path.exists(template_path):
        print(f'エラー: テンプレートが見つかりません → {template_path}')
        sys.exit(1)
    with open(template_path, encoding='utf-8') as f:
        html = f.read()

    meta = {}
    segments = []
    scroll_depths = []

    # ── XLSX（KO形式）処理
    # 来店結果報告データの自動読み込み
    visit_rate_data = None
    rate_file = find_data_file(script_dir, ['visit_rates.xlsx', '来店結果報告_平滑化版.xlsx', '報告結果_平滑化版.xlsx'], glob_pattern='*報告*平滑*.xlsx')
    if rate_file:
        visit_rate_data = load_visit_rates(rate_file)
        print(f'📈 来店結果報告データ読み込み済み: {len(visit_rate_data)}日分')
    else:
        print(f'⚠️  来店結果報告ファイルが見つかりません。固定値で代替します。')

    if args.xlsx:
        print(f'📊 KO XLSXを読み込み中: {args.xlsx}')
        meta, segments, age_segments = parse_ko_xlsx(args.xlsx, visit_rate_data=visit_rate_data)
        print(f'   → 店舗: {meta.get("store")}  送信ID: {meta.get("sendId")}')
        print(f'   → 送信成功: {meta.get("sent")}名  LP閲覧: {meta.get("lpViews")}名  来店: {meta.get("visits")}名')
        print(f'   → 離反セグメント: {len(segments)}  年代セグメント: {len(age_segments)}')
        html = inject_segments(html, segments)
        html = inject_age_segments(html, age_segments)

    # ── 設置台数・キャンペーン種別（インタラクティブ入力 or CLI引数）
    target_file  = find_data_file(script_dir, ['SMSターゲット.xlsx', 'SMS対象.xlsx', 'sms_targets.xlsx'])
    member_file  = find_data_file(script_dir, ['会員データ調査_想定台数100刻み_算出済.xlsx', '会員データ調査_想定人数100刻み_算出済.xlsx', 'member_data.xlsx'])
    campaign_targets = load_campaign_targets(target_file) if target_file else []

    if args.xlsx:
        # 設置台数：未指定ならプロンプト
        if not args.machines:
            try:
                val = input('🏪 店舗の設置台数を入力してください（例: 721）: ').strip()
                args.machines = int(val)
            except (ValueError, EOFError):
                print('⚠️  設置台数が無効のためシミュレーションをスキップします')

        # キャンペーン種別：未指定ならプロンプト
        max_segment   = None
        birthday_mode = False
        if campaign_targets and not args.campaign_type:
            print('\n📋 SMSの内容を選択してください:')
            for i, ct in enumerate(campaign_targets, 1):
                note = '（÷12補正あり）' if ct['birthday'] else ''
                print(f'  {i}. {ct["name"]}　→　〜{ct["max_seg"]}まで対象{note}')
            try:
                choice = int(input('番号を入力: ').strip()) - 1
                if 0 <= choice < len(campaign_targets):
                    selected = campaign_targets[choice]
                    args.campaign_type = selected['name']
                    max_segment        = selected['max_seg']
                    birthday_mode      = selected['birthday']
                    print(f'   → 選択: {args.campaign_type}（最大 {max_segment}）')
            except (ValueError, EOFError):
                print('⚠️  種別未選択のため全セグメントを対象とします')
        elif args.campaign_type:
            matched = next((ct for ct in campaign_targets if ct['name'] == args.campaign_type), None)
            if matched:
                max_segment   = matched['max_seg']
                birthday_mode = matched['birthday']
                print(f'📋 キャンペーン種別: {args.campaign_type}（最大 {max_segment}）')
            else:
                print(f'⚠️  キャンペーン種別 "{args.campaign_type}" が見つかりません。全セグメントを対象とします')

    if args.machines:
        meta['machines'] = args.machines
    if args.machines and args.xlsx and os.path.exists(member_file):
        print(f'🏪 設置台数 {args.machines}台 → 推定会員数を算出中...')
        member_estimates = load_member_estimates(member_file, args.machines)
        html = inject_unsent_segments(html, segments, member_estimates, visit_rate_data,
                                      meta.get('measurementDays', 0),
                                      max_segment=max_segment, birthday_mode=birthday_mode)
        print(f'   → 未送信層シミュレーションデータを注入完了')
        if args.campaign_type:
            html = inject_campaign_meta(html, args.campaign_type, max_segment or '')
    elif args.machines and not member_file:
        print(f'⚠️  会員データ調査ファイルが見つかりません')

    # ── Clarity CSV処理（Attention + Scroll 両対応）
    reach_map = {}
    if args.scroll_csv:
        print(f'📄 スクロールCSVを読み込み中: {args.scroll_csv}')
        reach_map = parse_scroll_csv(args.scroll_csv)
        print(f'   → 到達率データ {len(reach_map)}行')

    if args.csv:
        csv_type = detect_csv_type(args.csv)
        print(f'📄 注目割合CSVを読み込み中: {args.csv}')
        if csv_type in ('attention', 'clarity'):
            clarity_meta, scroll_depths = parse_clarity_csv(args.csv)
            print(f'   → Attentionデータ {len(scroll_depths)}行  総PV: {clarity_meta.get("pageViews")}')
            meta.update({k: v for k, v in clarity_meta.items() if k not in meta})
            if not meta.get('sendId') and clarity_meta.get('url'):
                m = re.search(r'/(\d+)$', clarity_meta['url'])
                if m: meta['sendId'] = m.group(1)
            # Scroll CSVから到達率を注入
            if reach_map:
                for d in scroll_depths:
                    d['reach'] = reach_map.get(d['depth'], None)
            html = inject_scroll_depths(html, scroll_depths)
        else:
            print('   ⚠️  未対応のCSV形式です')

    # ── SMS本文分析
    if args.xlsx and meta.get('smsText'):
        sms_analysis = analyze_sms(meta['smsText'], store_name=meta.get('store', ''))
        html = inject_sms_analysis(html, meta['smsText'], sms_analysis)
        score_label = {'good':'✅ 良好', 'caution':'⚠️ 要確認', 'review':'❗ 要改善'}.get(sms_analysis['score'], '')
        print(f'📝 SMS本文分析: {score_label}（{len(meta["smsText"])}字）')

    # ── KPI・ヘッダー注入（CSV読み込み後にまとめて実行）
    if args.xlsx:
        html = inject_kpi(html, meta)
        html = inject_header(html,
            store    = meta.get('store'),
            campaign = meta.get('campaign'),
            sendid   = meta.get('sendId'),
            dt       = meta.get('datetime'),
            axis     = '来店転換率 × LP支持率',
        )
    elif args.csv:
        html = inject_header(html, sendid=meta.get('sendId'))

    # ── 画像処理
    image_path = args.image
    if not image_path:
        base = os.path.splitext(args.xlsx or args.csv)[0]
        image_path = find_image(base)
        if image_path:
            print(f'🖼  画像を自動検出: {image_path}')
    if image_path:
        print(f'🖼  画像変換中...')
        html = inject_image(html, image_to_base64(image_path))
        print(f'   → 埋め込み完了')
    else:
        print('🖼  画像なし（同名の画像ファイルを置くか --image で指定）')

    # ── 出力
    out_path = os.path.join(script_dir, args.output)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'\n✅ レポート生成完了: {out_path}')
    print(f'   ファイルサイズ: {os.path.getsize(out_path) // 1024} KB')


if __name__ == '__main__':
    main()
